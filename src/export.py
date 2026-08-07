from __future__ import annotations

import csv
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable, Mapping

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


FIELDNAMES = [
    "title",
    "category",
    "price",
    "availability",
    "rating",
    "product_url",
    "image_url",
    "scrape_timestamp",
]
HEADERS = {
    "title": "Title",
    "category": "Category",
    "price": "Price",
    "availability": "Availability",
    "rating": "Rating",
    "product_url": "Product Page URL",
    "image_url": "Image URL",
    "scrape_timestamp": "Scrape Timestamp",
}


def write_rows_to_csv(rows: Iterable[Mapping[str, str]], output_file: Path) -> None:
    output_file.parent.mkdir(parents=True, exist_ok=True)

    with output_file.open("w", newline="", encoding="utf-8") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=FIELDNAMES)
        writer.writeheader()
        writer.writerows(rows)


def write_rows_to_excel(rows: Iterable[Mapping[str, str]], output_file: Path) -> None:
    output_file.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Books"

    worksheet.append([HEADERS[field] for field in FIELDNAMES])

    for row in rows:
        worksheet.append([format_excel_value(field, row.get(field, "")) for field in FIELDNAMES])

    style_header_row(worksheet)
    apply_column_formats(worksheet)
    apply_column_widths(worksheet)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    workbook.save(output_file)


def format_excel_value(field: str, value: str) -> object:
    if field == "price":
        return parse_price(value)

    if field == "rating":
        return int(value) if value else None

    if field == "scrape_timestamp" and value:
        parsed = datetime.fromisoformat(value)
        if parsed.tzinfo is not None:
            parsed = parsed.astimezone(timezone.utc).replace(tzinfo=None)
        return parsed

    return value


def parse_price(value: str) -> float | None:
    if not value:
        return None

    return float(value.replace("£", "").strip())


def style_header_row(worksheet) -> None:
    fill = PatternFill(fill_type="solid", fgColor="D9EAF7")

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = fill


def apply_column_formats(worksheet) -> None:
    column_formats = {
        "price": "£#,##0.00",
        "rating": "0",
        "scrape_timestamp": "yyyy-mm-dd hh:mm:ss",
    }

    for column_index, field in enumerate(FIELDNAMES, start=1):
        number_format = column_formats.get(field)
        if number_format is None:
            continue

        for row_index in range(2, worksheet.max_row + 1):
            worksheet.cell(row=row_index, column=column_index).number_format = number_format


def apply_column_widths(worksheet) -> None:
    min_widths = {
        "title": 24,
        "category": 16,
        "price": 12,
        "availability": 24,
        "rating": 10,
        "product_url": 42,
        "image_url": 42,
        "scrape_timestamp": 22,
    }
    max_widths = {
        "product_url": 70,
        "image_url": 70,
    }

    for column_index, field in enumerate(FIELDNAMES, start=1):
        column_letter = get_column_letter(column_index)
        header = HEADERS[field]
        longest_value = len(header)

        for cell in worksheet[column_letter]:
            if cell.value is not None:
                longest_value = max(longest_value, len(str(cell.value)))

        width = max(min_widths.get(field, 12), longest_value + 2)
        worksheet.column_dimensions[column_letter].width = min(width, max_widths.get(field, 45))
